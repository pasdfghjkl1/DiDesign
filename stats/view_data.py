# -*- coding: utf-8 -*-
"""
Скрипт для просмотра заявок и аналитики из Excel файла
Использование: python view_data.py --leads (для заявок)
              python view_data.py --analytics (для аналитики)
"""

import openpyxl
import glob
import argparse
from datetime import datetime

def find_excel_file():
    """Находит Excel файл в папке stats"""
    files = glob.glob(r'C:\Users\pa8hka\Desktop\site\stats\*.xlsx')
    if not files:
        raise FileNotFoundError("Excel file not found")
    return files[0]

def view_leads(limit=10):
    """Показывает последние заявки"""
    file_path = find_excel_file()
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Ищем лист
    sheet_name = None
    for variant in ["Заявки с сайта", "������ � �����"]:
        if variant in wb.sheetnames:
            sheet_name = variant
            break
    
    if not sheet_name:
        print("Sheet not found")
        return
    
    ws = wb[sheet_name]
    
    print("\n" + "="*80)
    print(f"LEADS (last {limit})".center(80))
    print("="*80)
    
    # Получаем заголовки
    headers = [cell.value for cell in ws[1]]
    
    # Получаем данные (последние N строк)
    start_row = max(2, ws.max_row - limit + 1)
    
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)]
        
        print(f"\n--- Lead #{row_data[0]} ---")
        print(f"Date: {row_data[1]} {row_data[2]}")
        print(f"Name: {row_data[3]}")
        print(f"Phone: {row_data[4]}")
        print(f"Email: {row_data[5]}")
        print(f"Type: {row_data[6]}")
        print(f"Message: {row_data[7]}")
        print(f"Page: {row_data[8]}")
        print(f"Status: {row_data[9]}")
        if row_data[10]:
            print(f"Comment: {row_data[10]}")
    
    print("\n" + "="*80)
    print(f"Total leads in database: {ws.max_row - 1}")
    print("="*80 + "\n")


def view_analytics(limit=10):
    """Показывает последнюю аналитику"""
    file_path = find_excel_file()
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Ищем лист
    sheet_name = None
    for variant in ["Аналитика сайта", "��������� �����"]:
        if variant in wb.sheetnames:
            sheet_name = variant
            break
    
    if not sheet_name:
        print("Sheet not found")
        return
    
    ws = wb[sheet_name]
    
    print("\n" + "="*80)
    print(f"ANALYTICS (last {limit} days)".center(80))
    print("="*80)
    
    # Получаем заголовки
    headers = [cell.value for cell in ws[1]]
    
    # Получаем данные
    start_row = max(2, ws.max_row - limit + 1)
    
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, len(headers) + 1)]
        
        print(f"\n--- {row_data[0]} ---")
        print(f"Unique visitors: {row_data[1]}")
        print(f"Total views: {row_data[2]}")
        print(f"Avg time: {row_data[3]} min")
        print(f"Pages: Home({row_data[4]}) Portfolio({row_data[5]}) Services({row_data[6]}) Drawings({row_data[7]}) About({row_data[8]})")
        print(f"Bounce rate: {row_data[9]}%")
        print(f"Source: {row_data[10]}")
    
    print("\n" + "="*80)
    print(f"Total days tracked: {ws.max_row - 1}")
    print("="*80 + "\n")

def main():
    parser = argparse.ArgumentParser(description='View data from Excel')
    parser.add_argument('--leads', action='store_true', help='Show leads')
    parser.add_argument('--analytics', action='store_true', help='Show analytics')
    parser.add_argument('--limit', type=int, default=10, help='Number of records to show')
    
    args = parser.parse_args()
    
    if args.leads:
        view_leads(args.limit)
    elif args.analytics:
        view_analytics(args.limit)
    else:
        print("Please specify --leads or --analytics")
        print("Example: python view_data.py --leads")

if __name__ == '__main__':
    main()
