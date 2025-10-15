# -*- coding: utf-8 -*-
"""
Скрипт для добавления заявки в Excel файл
Использование: python add_lead.py --name "Имя" --phone "+7..." --email "email@example.com"
"""

import openpyxl
from datetime import datetime
import glob
import sys
import argparse

def find_excel_file():
    """Находит Excel файл в папке stats"""
    files = glob.glob(r'C:\Users\pa8hka\Desktop\site\stats\*.xlsx')
    if not files:
        raise FileNotFoundError("Excel file not found in stats folder")
    return files[0]

def get_next_row_number(ws):
    """Получает номер следующей пустой строки"""
    return ws.max_row + 1

def add_lead(name, phone, email="", lead_type="Форма на сайте", 
             message="", page="", status="Новая", comment=""):
    """
    Добавляет заявку в Excel файл
    
    Параметры:
        name: Имя клиента
        phone: Телефон
        email: Email (необязательно)
        lead_type: Тип заявки
        message: Сообщение от клиента
        page: Страница, с которой пришла заявка
        status: Статус (Новая/В работе/Закрыта)
        comment: Комментарий менеджера
    """
    try:
        # Открываем файл
        file_path = find_excel_file()
        wb = openpyxl.load_workbook(file_path)
        
        # Ищем лист "Заявки с сайта"
        sheet_name = None
        for name_variant in ["Заявки с сайта", "������ � �����"]:
            if name_variant in wb.sheetnames:
                sheet_name = name_variant
                break
        
        if not sheet_name:
            raise ValueError("Sheet 'Заявки с сайта' not found")
        
        ws = wb[sheet_name]
        
        # Получаем номер следующей строки
        next_row = get_next_row_number(ws)
        row_number = next_row - 1  # Номер заявки
        
        # Текущие дата и время
        now = datetime.now()
        date_str = now.strftime("%d.%m.%Y")
        time_str = now.strftime("%H:%M:%S")
        
        # Структура: №, Дата, Время, Имя, Телефон, Email, 
        #           Тип заявки, Сообщение, Страница, Статус, Комментарий
        row_data = [
            row_number,      # № (колонка 1)
            date_str,        # Дата (колонка 2)
            time_str,        # Время (колонка 3)
            name,            # Имя (колонка 4)
            phone,           # Телефон (колонка 5)
            email,           # Email (колонка 6)
            lead_type,       # Тип заявки (колонка 7)
            message,         # Сообщение (колонка 8)
            page,            # Страница (колонка 9)
            status,          # Статус (колонка 10)
            comment          # Комментарий (колонка 11)
        ]
        
        # Записываем данные
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col, value=value)
        
        # Сохраняем файл
        wb.save(file_path)
        
        print(f"SUCCESS: Lead #{row_number} added")
        print(f"Name: {name}")
        print(f"Phone: {phone}")
        print(f"Email: {email}")
        print(f"Date: {date_str} {time_str}")
        
        return True
        
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        return False


def main():
    """Основная функция для CLI"""
    parser = argparse.ArgumentParser(description='Add lead to Excel file')
    parser.add_argument('--name', required=True, help='Client name')
    parser.add_argument('--phone', required=True, help='Client phone')
    parser.add_argument('--email', default='', help='Client email')
    parser.add_argument('--type', default='Форма на сайте', help='Lead type')
    parser.add_argument('--message', default='', help='Client message')
    parser.add_argument('--page', default='', help='Page URL')
    parser.add_argument('--status', default='Новая', help='Lead status')
    parser.add_argument('--comment', default='', help='Manager comment')
    
    args = parser.parse_args()
    
    success = add_lead(
        name=args.name,
        phone=args.phone,
        email=args.email,
        lead_type=args.type,
        message=args.message,
        page=args.page,
        status=args.status,
        comment=args.comment
    )
    
    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()
