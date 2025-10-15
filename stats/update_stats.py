# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import glob

# Находим файл
files = glob.glob(r'C:\Users\pa8hka\Desktop\site\stats\*.xlsx')
if not files:
    print("Файл не найден!")
    exit(1)

file_path = files[0]
print(f"Открываем файл: {file_path}")

# Открываем файл
wb = openpyxl.load_workbook(file_path)
print(f"Текущие листы: {wb.sheetnames}")

# ===== 1. СОЗДАЕМ ЛИСТ "Заявки с сайта" =====
sheet_name_leads = "Заявки с сайта"
if sheet_name_leads in wb.sheetnames:
    del wb[sheet_name_leads]
    print(f"Удален старый лист '{sheet_name_leads}'")

ws_leads = wb.create_sheet(sheet_name_leads)
print(f"Создан лист '{sheet_name_leads}'")

# Заголовки для заявок
headers_leads = [
    "№", "Дата", "Время", "Имя", "Телефон", "Email", 
    "Тип заявки", "Сообщение", "Страница", "Статус", "Комментарий"
]

# Добавляем заголовки с форматированием
header_fill = PatternFill(start_color="2C5530", end_color="2C5530", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers_leads, start=1):
    cell = ws_leads.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Устанавливаем ширину столбцов
column_widths_leads = {
    1: 5,   # №
    2: 12,  # Дата
    3: 10,  # Время
    4: 20,  # Имя
    5: 15,  # Телефон
    6: 25,  # Email
    7: 20,  # Тип заявки
    8: 40,  # Сообщение
    9: 20,  # Страница
    10: 15, # Статус
    11: 30  # Комментарий
}

for col, width in column_widths_leads.items():
    ws_leads.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

print(f"OK - Nastroyen list '{sheet_name_leads}'")

# ===== 2. СОЗДАЕМ ЛИСТ "Аналитика сайта" =====
sheet_name_analytics = "Аналитика сайта"
if sheet_name_analytics in wb.sheetnames:
    del wb[sheet_name_analytics]
    print(f"Удален старый лист '{sheet_name_analytics}'")

ws_analytics = wb.create_sheet(sheet_name_analytics)
print(f"Создан лист '{sheet_name_analytics}'")

# Заголовки для аналитики
headers_analytics = [
    "Дата", "Уникальных посетителей", "Всего просмотров", 
    "Среднее время на сайте (мин)", "Главная страница", 
    "Портфолио", "Услуги", "Чертежи", "О нас", 
    "Отказы (%)", "Источник трафика"
]

# Добавляем заголовки с форматированием
for col, header in enumerate(headers_analytics, start=1):
    cell = ws_analytics.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Устанавливаем ширину столбцов
column_widths_analytics = {
    1: 12,  # Дата
    2: 20,  # Уникальных посетителей
    3: 18,  # Всего просмотров
    4: 22,  # Среднее время
    5: 15,  # Главная
    6: 12,  # Портфолио
    7: 12,  # Услуги
    8: 12,  # Чертежи
    9: 10,  # О нас
    10: 12, # Отказы
    11: 20  # Источник
}

for col, width in column_widths_analytics.items():
    ws_analytics.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Добавляем пример данных для аналитики (текущий день)
today = datetime.now().strftime("%d.%m.%Y")
example_row = [today, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Прямой заход"]

for col, value in enumerate(example_row, start=1):
    cell = ws_analytics.cell(row=2, column=col, value=value)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

print(f"OK - Nastroyen list '{sheet_name_analytics}'")

# ===== 3. ПЕРЕУПОРЯДОЧИВАЕМ ЛИСТЫ =====
# Перемещаем новые листы в начало для удобства
sheets_order = [
    sheet_name_leads,
    sheet_name_analytics,
    "Data",
]

# Добавляем остальные листы которые уже были
for sheet in wb.sheetnames:
    if sheet not in sheets_order:
        sheets_order.append(sheet)

# Переупорядочиваем
wb._sheets = [wb[name] for name in sheets_order if name in wb.sheetnames]

print(f"Sheets order: {wb.sheetnames}")

# Сохраняем файл
wb.save(file_path)
print(f"\nOK - File saved: {file_path}")
print(f"OK - Added 2 new sheets:")
print(f"   - {sheet_name_leads}")
print(f"   - {sheet_name_analytics}")
