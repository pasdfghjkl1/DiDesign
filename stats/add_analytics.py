# -*- coding: utf-8 -*-
"""
Скрипт для добавления данных аналитики в Excel файл
Использование: python add_analytics.py --date "12.10.2025" --visitors 25 --views 120
"""

import openpyxl
from datetime import datetime
import glob
import sys
import argparse

def find_excel_file():
    """Находит Excel файл в папке stats"""
    files = glob.glob(r'C:\Users\pa8hka\Desktop\site\stats\*.xlsx')
    if not files:
        raise FileNotFoundError("Excel file not found in stats folder")
    return files[0]

def add_analytics(date_str, unique_visitors, total_views, avg_time,
                 page_home=0, page_portfolio=0, page_services=0, 
                 page_drawings=0, page_about=0, bounce_rate=0, 
                 traffic_source="Прямой заход"):
    """
    Добавляет данные аналитики в Excel файл
    
    Параметры:
        date_str: Дата в формате DD.MM.YYYY
        unique_visitors: Количество уникальных посетителей
        total_views: Всего просмотров страниц
        avg_time: Среднее время на сайте в минутах
        page_home: Просмотров главной страницы
        page_portfolio: Просмотров портфолио
        page_services: Просмотров услуг
        page_drawings: Просмотров чертежей
        page_about: Просмотров "О нас"
        bounce_rate: Процент отказов
        traffic_source: Источник трафика
    """
    try:
        # Открываем файл
        file_path = find_excel_file()
        wb = openpyxl.load_workbook(file_path)
        
        # Ищем лист "Аналитика сайта"
        sheet_name = None
        for name_variant in ["Аналитика сайта", "��������� �����"]:
            if name_variant in wb.sheetnames:
                sheet_name = name_variant
                break
        
        if not sheet_name:
            raise ValueError("Sheet 'Аналитика сайта' not found")
        
        ws = wb[sheet_name]
        
        # Проверяем, есть ли уже данные за эту дату
        date_found = False
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == date_str:
                # Обновляем существующую запись
                target_row = row
                date_found = True
                break
        
        if not date_found:
            # Создаем новую запись
            target_row = ws.max_row + 1
        
        # Структура: Дата, Уникальных посетителей, Всего просмотров, 
        #           Среднее время, Главная, Портфолио, Услуги, 
        #           Чертежи, О нас, Отказы (%), Источник трафика
        row_data = [
            date_str,           # Дата (колонка 1)
            unique_visitors,    # Уникальных посетителей (колонка 2)
            total_views,        # Всего просмотров (колонка 3)
            avg_time,           # Среднее время (колонка 4)
            page_home,          # Главная (колонка 5)
            page_portfolio,     # Портфолио (колонка 6)
            page_services,      # Услуги (колонка 7)
            page_drawings,      # Чертежи (колонка 8)
            page_about,         # О нас (колонка 9)
            bounce_rate,        # Отказы % (колонка 10)
            traffic_source      # Источник (колонка 11)
        ]
        
        # Записываем данные
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=target_row, column=col, value=value)
        
        # Сохраняем файл
        wb.save(file_path)
        
        action = "updated" if date_found else "added"
        print(f"SUCCESS: Analytics {action} for {date_str}")
        print(f"Visitors: {unique_visitors}, Views: {total_views}, Avg time: {avg_time} min")
        
        return True
        
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        return False

def main():
    """Основная функция для CLI"""
    parser = argparse.ArgumentParser(description='Add analytics to Excel file')
    parser.add_argument('--date', required=True, help='Date (DD.MM.YYYY)')
    parser.add_argument('--visitors', type=int, required=True, help='Unique visitors')
    parser.add_argument('--views', type=int, required=True, help='Total page views')
    parser.add_argument('--avgtime', type=float, default=0, help='Average time on site (minutes)')
    parser.add_argument('--home', type=int, default=0, help='Home page views')
    parser.add_argument('--portfolio', type=int, default=0, help='Portfolio page views')
    parser.add_argument('--services', type=int, default=0, help='Services page views')
    parser.add_argument('--drawings', type=int, default=0, help='Drawings page views')
    parser.add_argument('--about', type=int, default=0, help='About page views')
    parser.add_argument('--bounce', type=float, default=0, help='Bounce rate (%)')
    parser.add_argument('--source', default='Прямой заход', help='Traffic source')
    
    args = parser.parse_args()
    
    success = add_analytics(
        date_str=args.date,
        unique_visitors=args.visitors,
        total_views=args.views,
        avg_time=args.avgtime,
        page_home=args.home,
        page_portfolio=args.portfolio,
        page_services=args.services,
        page_drawings=args.drawings,
        page_about=args.about,
        bounce_rate=args.bounce,
        traffic_source=args.source
    )
    
    sys.exit(0 if success else 1)

if __name__ == '__main__':
    main()
